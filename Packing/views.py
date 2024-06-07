from django.shortcuts import get_object_or_404, redirect, render, HttpResponse, HttpResponseRedirect
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.management import call_command, CommandError
from django.http import JsonResponse
from django.forms import inlineformset_factory
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import timedelta, datetime, date
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from django.contrib import messages
from django.db.models import Count, Sum, F,DateField
from django.db.models.functions import Concat
import json
import os
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from SR_Plant_I.models import CustomUser
from django.db.models.functions import Cast
from django.db.models import Max, Subquery, OuterRef, Q
import pandas as pd


from .models import (branddetails, oilcategorydetails, skunamedetails, 
                    PrintingRollBatch, PrintingRollDetail, 
                    ProductionRollDetails,DispatchOpendingClosingStockDetails,
                    OilPumpingDetails, 
                    ChangeLog,
                    DailyPouchCuttingDetails,
                    ManualLeakChangeManpower,ManualLeakChangeRollPouchFS,
                    ExpVsActDetails,
                    PPSRDetails,
                    SRDailyStockDetails,
                    GodownDetails,
                    DispatchReq,
                    )
from .forms import (branddetailsform,
                    prinitingrolldetailsform,RollDetailsFormset, 
                    ProductionRollDetailsForm,
                    OilPumpingDetailsForm,
                    DailyPouchCuttingDetailsForm,
                    ManualLeakChangeManpowerForm,ManualLeakChangeRollPouchFSForm,
                    ExpVsActDetailsForm,
                    DispatchStockUploadForm,
                    #PPSRDetailsFormSet,
                    PPSRDetailsForm,SRDailyStockDetailsForm,
                    DispatchReqForm,
                    )
@method_decorator(login_required, name='dispatch')
class home(View):
    def get(self, request):
        breadcrumbs = [
            {'label': 'Home', 'url': '#'}
        ]
        context = {'breadcrumbs': breadcrumbs}

        try:
            userid = request.session['userdata']
            if userid:
                userdetails = CustomUser.objects.get(id=userid)
                ##print(userdetails.Pack_dashboard_access)
                context['user'] = userdetails
            else:
                # Handle case where user ID is not in session
                #print('No user ID in session')
                return redirect(reverse_lazy('login'))
        except KeyError:
            # Handle KeyError when 'userdata_pk' is not in session
            #print('No user ID in session (KeyError)')
            return redirect(reverse_lazy('login'))
        except CustomUser.DoesNotExist:
            # Handle case where the user does not exist
            #print('User not found')
            return redirect(reverse_lazy('login'))

        return render(request, 'Packing/dashboard.html', context)
  
@method_decorator(login_required,name='dispatch')   
class skudetails(ListView):
    model = branddetails
    template_name = 'Packing/skulist.html' 
    paginate_by = 10
    def paginate_custom_queryset(self, queryset, page_size,page_param):
        paginator = Paginator(queryset, page_size)
        page = self.request.GET.get(page_param)
        try:
            paginated_queryset = paginator.page(page)
        except PageNotAnInteger:
            paginated_queryset = paginator.page(1)
        except EmptyPage:
            paginated_queryset = paginator.page(paginator.num_pages)
        return paginated_queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'label': 'Home', 'url': '/Packing'},
            {'label': 'SKU\'s', 'url': None},
            {'label': 'SKU List', 'url': None},
        ]
        
        try:
            userid = self.request.session.get('userdata')
            #print('User ID' , userid)
            if userid:
                userdetails = CustomUser.objects.get(id=userid)
                #print(userdetails.Pack_sku_access)
                context['user'] = userdetails
            else:
                context['user'] = None              
        except KeyError:
            return reverse_lazy('login')
        except CustomUser.DoesNotExist:
            return reverse_lazy('login')
        # Paginate oilcategorydetails
        oilcategorylists = oilcategorydetails.objects.all()
        context['oilcategorylists'] = self.paginate_custom_queryset(oilcategorylists, self.paginate_by, 'page_oil')

        # Paginate skunamedetails
        skunamelists = skunamedetails.objects.all()
        #print(skunamelists)
        context['skunamelists'] = self.paginate_custom_queryset(skunamelists, self.paginate_by, 'page_sku')

        # Paginate branddetails
        brandlists = branddetails.objects.all()
        ##print("brandlists", brandlists)
        context['brandlists'] = self.paginate_custom_queryset(brandlists, self.paginate_by, 'page_brand')

        context['brandcount'] = brandlists.count()
        context['categorycount'] = oilcategorylists.count()
        context['skucount'] = skunamelists.count()

        return context
@method_decorator(login_required, name='dispatch')  
class filmrolltable(View):
    def get(self, request):
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Printing', 'url': None},
        {'label': 'Film Roll Table', 'url': None},  # Assuming current page doesn't have a URL
        ] 
        context = {'breadcrumbs': breadcrumbs}  
        
        userid = request.session['userdata']
        try:
            userdetails = CustomUser.objects.get(id=userid)
            #print(userdetails.Pack_printing_access)
            context['user'] = userdetails
        except:
            context['user'] = None
            
        query = PrintingRollDetail.objects.all().order_by('-updatedat')  # Order by the updated at in descending order
        paginator = Paginator(query, 10)  # Paginate by 10 items per page
        page = request.GET.get('page')
        try:
            datatable = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            datatable = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            datatable = paginator.page(paginator.num_pages)
        context['datatable'] = datatable     
        
        return render(request, 'Packing/printingfilmrolltable.html',context)    
class filmrolltableUpdateView(CreateView):
   pass
class filmrolltableDeleteView(DeleteView):
    pass
@login_required()
def filmrollentry(request):
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Printing', 'url': None},
        {'label': 'Film Roll Entry', 'url': None},  # Assuming current page doesn't have a URL
    ] 
    if request.method == 'POST':
        userid = request.session['userdata']
        try:
            userdetails = CustomUser.objects.get(id=userid)
            #print(userdetails.Pack_printing_access)            
        except:
            userdetails = None
        printing_rolldetails_form = prinitingrolldetailsform(request.POST)
        roll_detail_formset = RollDetailsFormset(request.POST)
        if roll_detail_formset.is_valid():
            if roll_detail_formset.has_changed():  # Check if there are any changes in the formset
                if printing_rolldetails_form.is_valid():
                    batch = printing_rolldetails_form.save()
                    roll_detail_formset.instance = batch
                    roll_detail_formset.save()
                    messages.success(request, 'Your details have been submitted successfully!')
                    return redirect('printingfilmrolltable')
                else:
                    messages.error(request, 'There was an error submitting the printing details.')
            else:
                messages.warning(request, 'No roll details were provided.')
        else:
            messages.error(request, 'There was an error submitting the formset.')
        
    else:
        userid = request.session['userdata']
        try:
            userdetails = CustomUser.objects.get(id=userid)
            #print(userdetails.Pack_printing_access)
        except:
            userdetails
        printing_rolldetails_form = prinitingrolldetailsform()
        roll_detail_formset = RollDetailsFormset()
    return render(request, 'Packing/printingfilmrollentry.html',{'printingrolldetailsform1':printing_rolldetails_form, 'rolldetailsformset':roll_detail_formset, 'breadcrumbs': breadcrumbs, 'user':userdetails},)#
def printingfilmrollavailability(request):
    
    userid = request.session['userdata']
    userdetails = CustomUser.objects.get(id=userid)    
    return render(request,'Packing/printingfilmrolltableavail.html',{'user':userdetails})    
def filmrollavilcheck(request):
    if request.method == 'GET' and 'rollno' in request.GET:
        rollno = request.GET['rollno']
        
        if rollno:        
            # Check if the roll number exists in the database
            if PrintingRollDetail.objects.filter(filmrollno=rollno).exists():
                result = {'message': 'Roll number already available'}
            else:
                result = {'message': 'New Roll number'}
            return JsonResponse(result)
        else:
            result = {'message':'Enter valid Roll number'}
            return JsonResponse(result)    
    else:
        # If roll_number is not provided or method is not GET, return error
        result = {'message':'Enter valid Roll number'}
        return JsonResponse(result)
    
@method_decorator(login_required,name='dispatch')        
class productionrolltableListView(ListView):
    model = ProductionRollDetails
    template_name = 'Packing/productionrolltable.html'
    context_object_name = 'datatable'
    paginate_by = 10
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/MIS/productionrolltable'},
        {'label': 'Production Table', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        
        userid = self.request.session.get('userdata')
        #print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
class productionrolltableCreateView(CreateView):
    model = ProductionRollDetails
    form_class = ProductionRollDetailsForm
    template_name = 'Packing/productionrolltableentry.html'
    success_url = reverse_lazy('productionrolltable')
    
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/MIS/productionrolltable'},
        {'label': 'Production Table Add', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def form_valid(self, form):
        messages.success(self.request, 'Production roll details have been successfully created.')
        return super().form_valid(form)
        
    pass

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        
        userid = self.request.session.get('userdata')
        #print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
class productionrolltableUpdateView(UpdateView):
    model = ProductionRollDetails
    form_class = ProductionRollDetailsForm
    template_name = 'Packing/productionrolltableedit.html'
    success_url = reverse_lazy('productionrolltable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/MIS/productionrolltable'},
        {'label': 'Film roll Update', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        
        userid = self.request.session.get('userdata')
        #print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Production roll details have been successfully Updated.')
        return super().form_valid(form)
class productionrolltableDeleteView(DeleteView):
    model = ProductionRollDetails
    template_name = 'Packing/productionrolltabledelete.html'
    success_url = reverse_lazy('productionrolltable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/MIS/productionrolltable'},
        {'label': 'Film Roll Delete', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        
        userid = self.request.session.get('userdata')
        #print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        
        return context
    
    def form_valid(self, form):
        messages.info(self.request, "Production roll details have been successfully Deleted.")
        return super().form_valid(form)
class dispatchstocktableListView(ListView):
    model = DispatchOpendingClosingStockDetails
    template_name = 'Packing/dispatchstocktable.html'
    context_object_name = 'datatable'
    paginate_by = 20
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/MIS/productionrolltable'},
        {'label': 'Dispatch Stock', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        context['form'] = DispatchStockUploadForm()  # Add upload form to context
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def post(self, request, *args, **kwargs):
            #print("post triggered")
            if 'file' in request.FILES:
                file = request.FILES['file']
                #print("File received:", file)
                #print("File name:", file.name)
            if not file.name.endswith('.xlsx'):
                messages.error(request, 'Unsupported file type. Please upload an Excel file') #with .xlsx extension.
                return redirect('dispatchstocktable')
            
            form = DispatchStockUploadForm(request.POST, request.FILES)
            #print("Form is valid:", form.is_valid())
            #print(form)
            if form.is_valid():
                try :
                    #print("File received:", request.FILES['file'])
                    file_path = self.handle_uploaded_file(request.FILES['file'])
                    call_command('import_excel_dispatchstock', file_path)
                    os.remove(file_path)  # Remove temporary file after import
                    messages.success(self.request, 'File have been successfully upload.')
                    
                except CommandError as e:
                    messages.error(request, str(e))
                except Exception as e:
                    messages.error(request, f"An unexpected error occurred: {str(e)}")
                return redirect('dispatchstocktable')  # Redirect to dispatch stock list view    
            else:
                #print("form not valid")
                messages.error(self.reqeust, "File not support")
                context = self.get_context_data(**kwargs)
                context['form'] = form
                return render(request, self.template_name, context)

    def handle_uploaded_file(self, file):
        #print("exe start")
        temp_dir = os.path.join(settings.BASE_DIR, 'temp_files')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        file_path = os.path.join(temp_dir, file.name)
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        return file_path


@method_decorator(login_required,name='dispatch')  
class oilpumpListView(ListView):
    model = OilPumpingDetails
    template_name = 'Packing/oilpumpingtable.html'
    context_object_name = 'datatable'
    paginate_by = 10
    
    
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Oil Pumping', 'url': '/Packing/oilpumpingtable'},
        {'label': 'Oil Pumping Table', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    pass   
class oilpumpCreateView(CreateView):
    model = OilPumpingDetails
    form_class = OilPumpingDetailsForm
    template_name = 'Packing/oilpumpingentry.html'
    success_url = reverse_lazy('oilpumpingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Oil Pumping', 'url': '/Packing/oilpumpingtable'},
        {'label': 'Oil Pumping Table - Add', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def form_valid(self, form):
        messages.success(self.request, 'Oil Pumping details have been successfully created.')
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    pass 
class oilpumpUpdateView(UpdateView):
    model = OilPumpingDetails
    form_class = OilPumpingDetailsForm
    template_name = 'Packing/oilpumpingtableedit.html'
    success_url = reverse_lazy('oilpumpingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Oil Pumping', 'url': '/Packing/oilpumpingtable'},
        {'label': 'Oil Pumping Edit', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Oil Pumping details have been successfully Updated.')
        return super().form_valid(form)
    
    pass 
class oilpumpDeleteView(DeleteView):
    model = OilPumpingDetails
    template_name = 'Packing/oilpumpingtabledelete.html'
    success_url = reverse_lazy('oilpumpingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Oil Pumping', 'url': '/Packing/oilpumpingtable'},
        {'label': 'Oil Pumping Delete', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        return context
    
    def form_valid(self, form):
        messages.info(self.request, "Oil Pumping details have been successfully Deleted.")
        return super().form_valid(form)
    
    pass 

@method_decorator(login_required,name='dispatch')
class DailyPouchCuttingDetailsListView(ListView):
    model = DailyPouchCuttingDetails
    template_name = 'Packing/dailypouchcuttingtable.html'
    context_object_name = 'datatable'
    paginate_by = 10
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Pouch Cutting', 'url': '/Packing/pouchcuttingtable'},
        {'label': 'Pouch Cutting Table', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    pass
class DailyPouchCuttingDetailsCreateView(CreateView):
    model = DailyPouchCuttingDetails
    form_class = DailyPouchCuttingDetailsForm
    template_name = 'Packing/dailypouchcuttingentry.html'
    success_url = reverse_lazy('pouchcuttingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Pouch Cutting', 'url': '/Packing/pouchcuttingtable'},
        {'label': 'Pouch Cutting Add', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Pouch Cutting details have been successfully created.')
        return super().form_valid(form)
class DailyPouchCuttingDetailsUpdateView(UpdateView):
    model = DailyPouchCuttingDetails
    form_class = DailyPouchCuttingDetailsForm
    template_name = 'Packing/dailypouchcuttingentry.html'
    success_url = reverse_lazy('pouchcuttingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Pouch Cutting', 'url': '/Packing/pouchcuttingtable'},
        {'label': 'Pouch Cutting Update', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def form_valid(self, form):
        messages.success(self.request, "Pouch Cutting details have been successfully Updated.")
        return super().form_valid(form)
    
    pass
class DailyPouchCuttingDetailsDeleteView(DeleteView):
    model = DailyPouchCuttingDetails
    template_name = 'Packing/dailypouchcuttingtabledelete.html'
    success_url = reverse_lazy('pouchcuttingtable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Pouch Cutting', 'url': '/Packing/pouchcuttingtable'},
        {'label': 'Pouch Cutting Delete', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def form_valid(self, form):
        messages.info(self.request, "Pouch Cutting details have been successfully Deleted.")
        return super().form_valid(form)
    pass   

@method_decorator(login_required, name='dispatch')
class ManualLeakChangeCreateView(CreateView):
    model = ManualLeakChangeManpower
    form_class = ManualLeakChangeManpowerForm
    template_name = 'Packing/manualleakchangetableentry.html'
    success_url = reverse_lazy('manualleakchangetable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Manual Leak Change', 'url': '/Packing/manualleakchangetable'},
        {'label': 'Leak Change Table Add', 'url': None},  # Assuming current page doesn't have a URL
        ]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['child_formset'] = inlineformset_factory(ManualLeakChangeManpower, ManualLeakChangeRollPouchFS, form=ManualLeakChangeRollPouchFSForm, extra=10)(self.request.POST)
        else:
            data['child_formset'] = inlineformset_factory(ManualLeakChangeManpower, ManualLeakChangeRollPouchFS, form=ManualLeakChangeRollPouchFSForm, extra=10)()
        data['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        data['user'] = userdetails
        return data
    
   
    def form_valid(self, form):
        context = self.get_context_data()
        child_formset = context['child_formset']
        if child_formset.is_valid():
            self.object = form.save()
            child_formset.instance = self.object
            child_formset.save()
            messages.success(self.request, 'Your details have been submitted successfully!')
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))
class ManualLeakChangeListView(ListView):
    model = ManualLeakChangeRollPouchFS
    template_name = 'Packing/manualleakchangetable.html'
    context_object_name = 'datatable'
    paginate_by = 10
    
    # obj = ManualLeakChangeRollPouchFS.objects.first()  # or filter by specific criteria

    # # Access the related PrintingRollDetail object
    # printing_roll_detail = obj.rollno

    # # Access the related PrintingRollBatch object
    # printing_roll_batch = printing_roll_detail.printingrollbatch

    # # Access the skuname
    # skuname = printing_roll_batch.skuname

    # # Print the skuname (or use it as needed)
    # #print(skuname)
    
    
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Manual Leak Change', 'url': '/Packing/manualleakchangetable'},
        {'label': 'Leak Change Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
    
    def get_queryset(self):
        # Order by updatedat field in descending order
        return ManualLeakChangeRollPouchFS.objects.all().order_by('-updatedat') 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
class ManualLeakChangeUpdateView(UpdateView):
    model = ManualLeakChangeManpower
    form_class = ManualLeakChangeManpowerForm
    template_name = 'Packing/manualleakchangetableedit.html'
    success_url = reverse_lazy('packinghome')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Manual Leak Change', 'url': '/Packing/manualleakchangetable'},
        {'label': 'Leak Change Table Update', 'url': None},  # Assuming current page doesn't have a URL
        ]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['child_formset'] = inlineformset_factory(ManualLeakChangeManpower, ManualLeakChangeRollPouchFS, form=ManualLeakChangeRollPouchFSForm, extra=10)(self.request.POST)
        else:
            data['child_formset'] = inlineformset_factory(ManualLeakChangeManpower, ManualLeakChangeRollPouchFS, form=ManualLeakChangeRollPouchFSForm, extra=10)()
        data['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        data['user'] = userdetails
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        child_formset = context['child_formset']
        if child_formset.is_valid():
            self.object = form.save()
            child_formset.instance = self.object
            child_formset.save()
            messages.success(self.request, 'Your details have been updated successfully!')
            return super().form_valid(form)
        else:
            messages.error(self.request, 'Your details not updated!')
            return self.render_to_response(self.get_context_data(form=form))
class ManualLeakChangeDeleteView(DeleteView):
    model = ManualLeakChangeManpower
    template_name = 'Packing/brandnameedit.html'
    success_url = reverse_lazy('manualleakchangetable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Manual Leak Change', 'url': '/Packing/manualleakchangetable'},
        {'label': 'Leak Change Table Update', 'url': None},  # Assuming current page doesn't have a URL
        ]
    
    def form_valid(self, form):
        messages.info(self.request, "Pouch Leak Change details have been successfully Deleted.")
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        context =  super().get_context_data(**kwargs)
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        context['breadcrumbs'] = self.breadcrumbs
def autocomplete_skuname(request):
    
    if 'term' in request.GET:
        return JsonResponse({'hi':'hi'})
        #print('data received')
        qs = skunamedetails.objects.filter(name__icontains='su')
        names = list()
        for item in qs:
            names.append(item.name)
        return JsonResponse(names, safe=False)
    return JsonResponse([], safe=False)

logger = logging.getLogger(__name__)
def toggle(request):
    try:
        btnradio = request.GET.get('btnradio')
        filter_option,dbname = btnradio.split(',')
        ##print("option",filter_option)
        ##print("dbname",dbname)
        if dbname == "ProductionRollDetails":   
            today = datetime.now().date()
            if filter_option == 'all':
                data = ProductionRollDetails.objects.all().order_by('-updatedat')
            elif filter_option == 'today':
                data = ProductionRollDetails.objects.filter(runningdate=today)
            elif filter_option == '7d':
                seven_days_ago = today - timedelta(days=7)
                ##print(seven_days_ago)
                data = ProductionRollDetails.objects.filter(runningdate__gte=seven_days_ago)
            elif filter_option == 'last_month':
                last_month_start = today.replace(day=1) - timedelta(days=1)
                last_month_end = today.replace(day=1) - timedelta(days=1)
                data = ProductionRollDetails.objects.filter(runningdate__range=[last_month_start, last_month_end])
            elif filter_option == 'last_3_months':
                three_months_ago = today.replace(day=1) - timedelta(days=1)
                last_three_months_start = three_months_ago - timedelta(days=90)
                data = ProductionRollDetails.objects.filter(runningdate__gte=last_three_months_start)
            elif filter_option == 'last_year':
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                data = ProductionRollDetails.objects.filter(runningdate__range=[last_year_start, last_year_end])
            else:
                # Handle invalid filter options
                data = ProductionRollDetails.objects.none()            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')
            return render(request, 'Packing/productionrolltable_data.html', {'datatable': data})
        
        if dbname == "PrintingRollDetail":   
            today = datetime.now().date()
            if filter_option == 'all':
                data = PrintingRollDetail.objects.all()
            elif filter_option == 'today':
                # batchdate = PrintingRollBatch.objects.filter(date=today)
                # data = PrintingRollDetail.objects.filter(printingrollbatch=batchdate)
                 # Filter PrintingRollBatch for the date range
                batchdate = PrintingRollBatch.objects.filter(date__range=[today, today])
                        
                # Get all instances if any
                batch_instances = batchdate.all()
                
                data = PrintingRollDetail.objects.filter(printingrollbatch__in=batch_instances)
            elif filter_option == '7d':
                seven_days_ago = today - timedelta(days=7)
                batchdate = PrintingRollBatch.objects.filter(date__gte=seven_days_ago)
                ##print(batchdate)
                data = PrintingRollDetail.objects.filter(printingrollbatch__in=batchdate)
            elif filter_option == 'last_month':
                # last_month_start = today.replace(day=1) - timedelta(days=1)
                # last_month_end = today.replace(day=1) - timedelta(days=1)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                batchdate = PrintingRollBatch.objects.filter(date__range=[last_month_start, last_month_end])
                ##print(batchdate)
                data = PrintingRollDetail.objects.filter(printingrollbatch__in=batchdate)
            elif filter_option == 'last_3_months':
                # three_months_ago = today.replace(day=1) - timedelta(days=1)
                # last_three_months_start = three_months_ago - timedelta(days=90)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_3_months_start = last_month_end.replace(day=1) - timedelta(days=2*30)  # Assuming 30 days per month
                last_3_months_end = last_month_end
    
                # Filter PrintingRollBatch objects for the last 3 months
                batchdate = PrintingRollBatch.objects.filter(date__range=[last_3_months_start, last_3_months_end])
                # #print(last_3_months_start)
                # #print(last_3_months_end)
                
                # Filter PrintingRollDetail objects based on batchdate
                data = PrintingRollDetail.objects.filter(printingrollbatch__in=batchdate)
            elif filter_option == 'last_year':
                # last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                # last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                # data = PrintingRollDetail.objects.filter(date__range=[last_year_start, last_year_end])
                today = date.today()
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = last_year_start.replace(month=12, day=31)
                
                # Filter PrintingRollBatch objects for the last year
                batchdate = PrintingRollBatch.objects.filter(date__range=[last_year_start, last_year_end])
                ##print(batchdate)
                
                # Filter PrintingRollDetail objects based on batchdate
                data = PrintingRollDetail.objects.filter(printingrollbatch__in=batchdate)
            else:
                # Handle invalid filter options
                data = PrintingRollDetail.objects.none()
            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')

            return render(request, 'Packing/printingfilmrolltable_data.html', {'datatable': data})
        
        if dbname == "OilPumpingDetails":   
            today = datetime.now().date()
            if filter_option == 'all':
                data = OilPumpingDetails.objects.all()
            elif filter_option == 'today':
                
                data = OilPumpingDetails.objects.filter(date=today)
            elif filter_option == '7d':
                seven_days_ago = today - timedelta(days=7)
                ##print(seven_days_ago)
                data = OilPumpingDetails.objects.filter(date__gte=seven_days_ago)
            elif filter_option == 'last_month':
                # last_month_start = today.replace(day=1) - timedelta(days=1)
                # last_month_end = today.replace(day=1) - timedelta(days=1)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                batchdate = OilPumpingDetails.objects.filter(date__range=[last_month_start, last_month_end])
                ##print(last_month_end)
                ##print(last_month_start)
                data = OilPumpingDetails.objects.filter(date__range=[last_month_start, last_month_end])
            elif filter_option == 'last_3_months':
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_3_months_start = last_month_end.replace(day=1) - timedelta(days=2*30)  # Assuming 30 days per month
                last_3_months_end = last_month_end
                #print(last_3_months_start)
                #print(last_3_months_end)
                data = OilPumpingDetails.objects.filter(date__gte=last_3_months_start)
            elif filter_option == 'last_year':
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                data = OilPumpingDetails.objects.filter(date__range=[last_year_start, last_year_end])
            else:
                # Handle invalid filter options
                data = OilPumpingDetails.objects.none()            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')
            return render(request, 'Packing/oilpumpingtable_data.html', {'datatable': data})
        
        if dbname == "DailyPouchCuttingDetails":   
            today = datetime.now().date()
            if filter_option == 'all':
                data = DailyPouchCuttingDetails.objects.all()
            elif filter_option == 'today':
                
                data = DailyPouchCuttingDetails.objects.filter(date=today)
            elif filter_option == '7d':
                seven_days_ago = today - timedelta(days=7)
                ##print(seven_days_ago)
                data = DailyPouchCuttingDetails.objects.filter(date__gte=seven_days_ago)
            elif filter_option == 'last_month':
                # last_month_start = today.replace(day=1) - timedelta(days=1)
                # last_month_end = today.replace(day=1) - timedelta(days=1)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                batchdate = DailyPouchCuttingDetails.objects.filter(date__range=[last_month_start, last_month_end])
                ##print(last_month_end)
                ##print(last_month_start)
                data = DailyPouchCuttingDetails.objects.filter(date__range=[last_month_start, last_month_end])
            elif filter_option == 'last_3_months':
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_3_months_start = last_month_end.replace(day=1) - timedelta(days=2*30)  # Assuming 30 days per month
                last_3_months_end = last_month_end
                # #print(last_3_months_start)
                # #print(last_3_months_end)
                data = DailyPouchCuttingDetails.objects.filter(date__gte=last_3_months_start)
            elif filter_option == 'last_year':
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                data = DailyPouchCuttingDetails.objects.filter(date__range=[last_year_start, last_year_end])
            else:
                # Handle invalid filter options
                data = DailyPouchCuttingDetails.objects.none()            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')
            return render(request, 'Packing/dailypouchcuttingtable_data.html', {'datatable': data})
        
        if dbname == "ManualLeakChangeRollPouchFS":   
            today = datetime.now().date()
            if filter_option == 'all':
                #print("hi")
                data = ManualLeakChangeRollPouchFS.objects.all()
            elif filter_option == 'today':
                manpowerdetails = ManualLeakChangeManpower.objects.filter(date=today)
                ##print(leakchangedate)
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower__in=manpowerdetails)
            elif filter_option == '7d':
                seven_days_ago = today - timedelta(days=7)
                manpowerdetails = ManualLeakChangeManpower.objects.filter(date__gte=seven_days_ago)
                ##print(batchdate)
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower__in=manpowerdetails)
            elif filter_option == 'last_month':
                # last_month_start = today.replace(day=1) - timedelta(days=1)
                # last_month_end = today.replace(day=1) - timedelta(days=1)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                manpowerdetails = ManualLeakChangeManpower.objects.filter(date__range=[last_month_start, last_month_end])
                ##print(batchdate)
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower__in=manpowerdetails)
            elif filter_option == 'last_3_months':
                # three_months_ago = today.replace(day=1) - timedelta(days=1)
                # last_three_months_start = three_months_ago - timedelta(days=90)
                today = date.today()
                last_month_end = today.replace(day=1) - timedelta(days=1)
                last_3_months_start = last_month_end.replace(day=1) - timedelta(days=2*30)  # Assuming 30 days per month
                last_3_months_end = last_month_end
    
                # Filter PrintingRollBatch objects for the last 3 months
                manpowerdetails = ManualLeakChangeManpower.objects.filter(date__range=[last_3_months_start, last_3_months_end])
                ##print(last_3_months_start)
                ##print(last_3_months_end)
                
                # Filter PrintingRollDetail objects based on batchdate
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower__in=manpowerdetails)
            elif filter_option == 'last_year':
                # last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                # last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                # data = PrintingRollDetail.objects.filter(date__range=[last_year_start, last_year_end])
                today = date.today()
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = last_year_start.replace(month=12, day=31)
                
                # Filter PrintingRollBatch objects for the last year
                manpowerdetails = ManualLeakChangeManpower.objects.filter(date__range=[last_year_start, last_year_end])
                ##print(batchdate)
                
                # Filter PrintingRollDetail objects based on batchdate
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower__in=manpowerdetails)
            else:
                # Handle invalid filter options
                data = ManualLeakChangeRollPouchFS.objects.none()
            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')

            return render(request, 'Packing/manualleakchangetable_data.html', {'datatable': data})
        if dbname == "ExpVsActDetails":   
            today = datetime.now().date()
            if filter_option == 'all':
                data = ExpVsActDetails.objects.all()
            elif filter_option == 'today':
                data = ExpVsActDetails.objects.filter(date=today)
            elif filter_option == 'thismonth':
                current_year = datetime.now().year
                current_month = datetime.now().month

                # Filter records from the current month
                data = ExpVsActDetails.objects.filter(date__year=current_year, date__month=current_month)
            elif filter_option == 'last_month':
                last_month_start = today.replace(day=1) - timedelta(days=1)
                last_month_end = today.replace(day=1) - timedelta(days=1)
                data = ExpVsActDetails.objects.filter(date__range=[last_month_start, last_month_end])
            elif filter_option == 'last_3_months':
                three_months_ago = today.replace(day=1) - timedelta(days=1)
                last_three_months_start = three_months_ago - timedelta(days=90)
                data = ExpVsActDetails.objects.filter(date__gte=last_three_months_start)
            elif filter_option == 'last_year':
                last_year_start = today.replace(year=today.year - 1, month=1, day=1)
                last_year_end = today.replace(year=today.year - 1, month=12, day=31)
                data = ExpVsActDetails.objects.filter(date__range=[last_year_start, last_year_end])
            else:
                # Handle invalid filter options
                data = ExpVsActDetails.objects.none()            
            # Log the data for debugging
            logger.info(f'Filtered data: {data}')
            return render(request, 'Packing/expvsacttable_data.html', {'datatable': data})
    except Exception as e:
        logger.error(f'Error in toggle view: {e}')
        # Handle exceptions, e.g., log the error
        return render(request, 'Packing/expvsacttable_data.html', {'datatable': []})
    
    
def download_printingrolltable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        # #print(from_date_str)
        # #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        # printingbatch = PrintingRollBatch.objects.filter(date__range=(from_date,to_date))        
        # queryset = PrintingRollDetail.objects.filter(printingrollbatch__in=printingbatch)
        # batchdate = PrintingRollBatch.objects.filter(date__range=[from_date, to_date])
        
        batchdate = PrintingRollBatch.objects.filter(date__range=[from_date,to_date])
        # #print(dir(batchdate))
        # skuname = skunamedetails.objects.filter(skuname = batchdate)
        ##print(batchdate)
        data = PrintingRollDetail.objects.filter(printingrollbatch__in=batchdate)
        ##print(data)
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Shift','SKU Name', 'MRP','Batch No','Roll No','Film Roll Date','Gross Wt','Net Wt','Operator Name']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:
                # #print(dir(obj))
                row = [obj.printingrollbatch.date, obj.printingrollbatch.shift.shifttype,obj.printingrollbatch.skuname.skuname, obj.printingrollbatch.mrp, obj.printingrollbatch.batch_no, obj.filmrollno, obj.filmrolldate, 
                    obj.grosswt,obj.netwt,obj.printingrollbatch.operatorname.operatorname]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Printingfilmroll_table.xlsx"'
            wb.save(response)
            
            return response
        else : 
            return HttpResponseRedirect('printingfilmrolltable')
       
    return HttpResponseRedirect('printingfilmrolltable')
def download_productionrolltable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        # #print(from_date_str)
        # #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        data = ProductionRollDetails.objects.filter(runningdate__range=[from_date,to_date])
        
        
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Shift','Machine No', 'SKU Name','Roll Type','Roll Start Time','Roll Stop Time',
                       'Batch No','MRP','Roll No','Operator Name','Pouch Count']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:
                # created_at = obj.createdat.replace(tzinfo=None) if obj.createdat else None
                updatedat = obj.updatedat.replace(tzinfo=None) if obj.updatedat else None
                # #print(dir(obj))
                
                def make_naive(dt):
                    return dt.replace(tzinfo=None) if dt else None

                runningrollstarttime = make_naive(obj.runningrollstarttime) if hasattr(obj, 'runningrollstarttime') else None
                runningrollstoptime = make_naive(obj.runningrollstoptime) if hasattr(obj,'runningrollstoptime') else None
                
                
                row = [obj.runningdate, obj.runningshift.shifttype,obj.runningmachine.machinename,obj.runningskuname.skuname, obj.runningrolltype.rolltype, runningrollstarttime,runningrollstoptime, 
                       obj.runningbatchno,obj.runningmrp,obj.runningrollno, obj.runningoperatorname.operatorname]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Productionfilroll_table.xlsx"'
            wb.save(response)
            
            return response
        else : 
            return HttpResponseRedirect('productionrolltable')
       
    return HttpResponseRedirect('productionrolltable')
def download_oilpumpingtable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        # #print(from_date_str)
        # #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        data = OilPumpingDetails.objects.filter(date__range=[from_date,to_date])
        
        
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Shift','Motor ON Time', 'Motor OFF Time','Main Tank','Sub Tank','Vitamin',
                       'TMPS','TBHQ','Operator Name','QC Name', 'Manager']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:                
                def make_naive(dt):
                    return dt.replace(tzinfo=None) if dt else None

                motorontime = make_naive(obj.motorontime) if hasattr(obj, 'motorontime') else None
                motorofftime = make_naive(obj.motorofftime) if hasattr(obj,'motorofftime') else None
                
                
                row = [obj.date, obj.shift.shifttype,motorontime,motorofftime, obj.maintank.maintankname, obj.subtank.subtankname,obj.vitaminunits.units, 
                       obj.tmpsunits.units,obj.tbhqunits.units,obj.operatorname.operatorname,obj.qcname.qcname,obj.manager.managername]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="OilPumping_table.xlsx"'
            wb.save(response)
            
            return response
        else : 
            return HttpResponseRedirect('oilpumpingtable')
       
    return HttpResponseRedirect('oilpumpingtable')
def download_pouchcuttingtable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        # #print(from_date_str)
        # #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        data = DailyPouchCuttingDetails.objects.filter(date__range=[from_date,to_date])
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Shift','Operator Name','SF Leak in MT', 'GN Leak in MT','GNR Leak in MT','RB Leak in MT','Palm Leak in MT',
                       'GIN Leak in MT']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:                
                def make_naive(dt):
                    return dt.replace(tzinfo=None) if dt else None
       
                row = [obj.date, obj.shift.shifttype,obj.operatorname.operatorname, obj.sfleakinmt, obj.gnleakinmt,obj.gnrleakinmt, 
                       obj.rbleakinmt,obj.palmleakinmt,obj.ginleakinmt]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="PouchCutting_table.xlsx"'
            wb.save(response)
            
            return response
        else : 
            return HttpResponseRedirect('download_pouchcuttingtable')
       
    return HttpResponseRedirect('download_pouchcuttingtable')
def download_manualleakchangetable(request):  
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        # Filter by date range
        manpowers = ManualLeakChangeManpower.objects.filter(date__range=[from_date, to_date])
        if manpowers.exists():
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Shift', 'SKU Name', 'Changed Box', 'Roll No', 'Total Pouch', 'Mistake', 'Total Manpower']
            ws.append(headers)
            
            column_widths = [len(header) for header in headers]
            
            for obj in manpowers:
                data = ManualLeakChangeRollPouchFS.objects.filter(manpower=obj)
                first_row = True
                for detail in data:
                    try:
                        skuname_detail = PrintingRollDetail.objects.get(filmrollno=detail.rollno)
                        skuname = skuname_detail.printingrollbatch.skuname.skuname
                    except PrintingRollDetail.DoesNotExist:
                        skuname = 'N/A'

                    totalmanpower = obj.hindimanpower + obj.ladiesmanpower
                    
                    rollno = detail.rollno.filmrollno
                    mistakename = detail.mistakename.mistakename
                    
                    row = [
                        obj.date,
                        obj.shift.shifttype,
                        skuname,
                        obj.changeddamagebox if first_row else '',
                        rollno,
                        detail.noofpouch,
                        mistakename,
                        totalmanpower
                    ]
                    ws.append(row)
                    
                    # Update the maximum column widths
                    for i, cell_value in enumerate(row):
                        column_widths[i] = max(column_widths[i], len(str(cell_value)))
            # Adjust the column widths based on the maximum lengths
            for i, column_width in enumerate(column_widths):
                column_letter = get_column_letter(i + 1)
                ws.column_dimensions[column_letter].width = column_width + 2  # Adding extra padding
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply border to all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ManualPouchLeakChange_table.xlsx"'
            wb.save(response)
            
            return response
        else:
            return HttpResponse("No data found for the given date range.")
       
    return HttpResponseRedirect('manualleakchangetable')

    # Return the HTML fragment as JSON response
    # return JsonResponse({'html': html_data})
    #return JsonResponse({'html': 'html_data'})
    # def get(self, request):
    #     query = ProductionRollDetails.objects.all().order_by('-updatedat')  # Order by the updated at in descending order
    #     breadcrumbs = [
    #     {'label': 'Home', 'url': '/Packing'},
    #     {'label': 'Production', 'url': None},
    #     {'label': 'Production Roll Table', 'url': None},  # Assuming current page doesn't have a URL
    # ] 
    #     context = {
    #         'breadcrumbs':breadcrumbs,
    #         'datatable': query
    #     }
            
    #     return render(request, 'Packing/productionrolltable.html',context)
def download_expvsacttable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        ##print("hit")
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        #print(from_date_str)
        #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        data = ExpVsActDetails.objects.filter(date__range=[from_date,to_date])
        ##print(data)
        
        
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'Expected Production Box','Actual Production Box', 'Different','Percentage %','Remarks']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:
                # created_at = obj.createdat.replace(tzinfo=None) if obj.createdat else None
                updatedat = obj.updatedat.replace(tzinfo=None) if obj.updatedat else None
                # #print(dir(obj))
                
                def make_naive(dt):
                    return dt.replace(tzinfo=None) if dt else None

                #date = make_naive(obj.date) if hasattr(obj, 'date') else None
                
                different = obj.actbox - obj.expbox       
                percentage = (obj.actbox / obj.expbox) * 100
                percentage = round(percentage, 2)        
                row = [obj.date, obj.expbox,obj.actbox,different,str(percentage) + '%',obj.remarks]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ExpectedVsActualProduction_table.xlsx"'
            wb.save(response)
            
            return response
        else :
            #print("Data not found") 
            return HttpResponseRedirect('expvsacttable')
       
    return HttpResponseRedirect('expvsacttable')
def download_dispatchstocktable(request):
    if request.method == 'GET' and 'fromdate' in request.GET and 'todate' in request.GET:
        ##print("hit")
        from_date_str = request.GET['fromdate']
        to_date_str = request.GET['todate']
        
        #print(from_date_str)
        #print(to_date_str)
        
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle invalid date format
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")
        
        data = DispatchOpendingClosingStockDetails.objects.filter(date__range=[from_date,to_date])
        ##print(data)
        
        
        
        if data.exists():
        
            wb = Workbook()
            ws = wb.active
            
            headers = ['Date', 'SKU Code','Category', 'SKU Name','Opening stock','Sales','Closing Stock','Production','No of Empty Box']  # Adjust according to your model fields
            ws.append(headers)
            
            for obj in data:
                # created_at = obj.createdat.replace(tzinfo=None) if obj.createdat else None
                updatedat = obj.updatedat.replace(tzinfo=None) if obj.updatedat else None
                # #print(dir(obj))
                
                def make_naive(dt):
                    return dt.replace(tzinfo=None) if dt else None

                #date = make_naive(obj.date) if hasattr(obj, 'date') else None  
                row = [obj.date, obj.skucode,obj.categoryname,obj.skuname,obj.openingstock,obj.sales,obj.closingstock,obj.production,obj.noofemptycottonbox]  # Adjust according to your model fields
                ws.append(row)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="DispatchSalesSheet_table.xlsx"'
            wb.save(response)
            
            return response
        else :
            #print("Data not found") 
            return HttpResponseRedirect('dispatchstocktable')
       
    return HttpResponseRedirect('dispatchstocktable')
    pass
def chart_data(request):
    # data = ProductionRollDetails.objects.values('runningdate').annotate(
    #     total_pouches=Sum('runningpouchcount'),
    #     machines=Count('runningmachine', distinct=True),
    #     operators=Count('runningoperatorname', distinct=True),
    #     machine_list=Concat(F('runningmachine__name'), output_field=models.CharField()),
    #     operator_list=Concat(F('runningoperatorname__name'), output_field=models.CharField())
    # ).order_by('runningdate')
    data = ProductionRollDetails.objects.values('runningdate').annotate(
        total_pouches=Sum('runningpouchcount'),
        machines=Count('runningmachine', distinct=True),
        operators=Count('runningoperatorname', distinct=False),
    ).order_by('runningdate')
    data1 = ProductionRollDetails.objects.values('runningdate').annotate(machines=Count('runningmachine', distinct=True))
    ##print("data Value : ", data)

    # Aggregate unique machine and operator names for each date
    for entry in data:
        entry['machine_list'] = list(ProductionRollDetails.objects.filter(runningdate=entry['runningdate'])
                                      .values_list('runningmachine__machinename', flat=True).distinct())
        entry['operator_list'] = list(ProductionRollDetails.objects.filter(runningdate=entry['runningdate'])
                                      .values_list('runningoperatorname__operatorname', flat=True).distinct())
    ##print(list(data))
    
    
    return JsonResponse(list(data), safe=False)

def chart_view(request):
    data = ProductionRollDetails.objects.values('runningdate').annotate(
        total_pouches=Sum('runningpouchcount'),
        machines=Count('runningmachine', distinct=True),
        operators=Count('runningoperatorname', distinct=True)
    ).order_by('runningdate')

    chart_data = [
        {
            'y': entry['runningdate'].strftime('%Y-%m-%d'),
            'a': entry['total_pouches'],
            'b': entry['machines'],
            'c': entry['operators']
        }
        for entry in data
    ]

    return render(request, 'Packing/chart.html', {'chart_data': json.dumps(chart_data)})

@method_decorator(login_required,name='dispatch')
class ExpVsActDetailsListView(ListView):
    model = ExpVsActDetails
    template_name = 'Packing/expvsacttable.html'
    context_object_name = 'datatable'
    paginate_by = 10
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Exp Vs Act', 'url': '/Packing/expvsacttable'},
        {'label': 'Exp Vs Act Table', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        
        return context
class ExpVsActDetailsCreateView(CreateView):
    model = ExpVsActDetails
    form_class = ExpVsActDetailsForm
    template_name = 'Packing/expvsacttableedit.html'
    success_url = reverse_lazy('expvsacttable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Exp Vs Act', 'url': '/Packing/expvsacttable'},
        {'label': 'Exp Vs Act Table Add', 'url': None},  # Assuming current page doesn't have a URL
        ]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Expected Vs Actual Production have been successfully created.')
        return super().form_valid(form)
class ExpVsActDetailsUpdateView(UpdateView):
    model = ExpVsActDetails
    form_class = ExpVsActDetailsForm
    template_name = 'Packing/expvsacttableedit.html'
    success_url = reverse_lazy('expvsacttable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Exp Vs Act', 'url': '/Packing/expvsacttable'},
        {'label': 'Exp Vs Act Table Edit', 'url': None},  # Assuming current page doesn't have a URL
        ]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Expected Vs Actual Production details have been successfully updated.')
        return super().form_valid(form)
class ExpVsActDetailsDeleteView(DeleteView):
    model = ExpVsActDetails
    template_name = 'Packing/expvsacttabledelete.html'
    success_url = reverse_lazy('expvsacttable')
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Exp Vs Act', 'url': '/Packing/expvsacttable'},
        {'label': 'Exp Vs Act Table Delete', 'url': None},  # Assuming current page doesn't have a URL
        ]
    def form_valid(self, form):
        messages.info(self.request, "Expected Vs Actual details have been successfully Deleted.")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        return context
class ExpVsActChartListView(ListView):
    model = ExpVsActDetails
    template_name='Packing/expvsactchart.html'
    context_object_name = 'datatable'
    
    def get_context_data(self, **kwargs):
        context =  super().get_context_data(**kwargs)
        userid = self.request.session.get('userdata')
        userdetails = CustomUser.objects.get(id=userid)
        # Retrieve data for the chart
        chart_data = list(ExpVsActDetails.objects.values('date', 'expbox', 'actbox'))
        
    #Month -------------------
        # Convert dates to string
        for item in chart_data:
            item['date'] = item['date'].strftime('%d-%m-%Y')
        
        # Organize data by month and calculate totals
        data_by_month = {}
        for item in chart_data:
            month = item['date'][3:10]  # Extract month and year (MM-YYYY)
            if month not in data_by_month:
                data_by_month[month] = {'exp_total': 0, 'act_total': 0}
            data_by_month[month]['exp_total'] += item['expbox']
            data_by_month[month]['act_total'] += item['actbox']
        
        # Convert data_by_month to a list of dictionaries
        chart_month_data = [{'month': month, 'exp_total': data['exp_total'], 'act_total': data['act_total']} for month, data in data_by_month.items()]
        context['chart_month'] = json.dumps(chart_month_data)  # Convert to JSON string for monthly totals
        
        
    #____________________________________________________________________________
    
    #Current month--------------------------------------------------------------
        # Get the current month and year
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Filter data for the current month
        chart_data = ExpVsActDetails.objects.filter(date__month=current_month, date__year=current_year).order_by('date')
        
        # Convert queryset to list of dictionaries
        chart_data_list = [{'date': item.date.strftime('%d-%m-%Y'), 'expbox': item.expbox, 'actbox': item.actbox} for item in chart_data]
        context['chart_day'] = json.dumps(chart_data_list)  # Convert to JSON string
    
        # print(chart_data)
        return context

@login_required()    
def ppsr_details_view(request):
    userid = request.seesion['userdata']
    userdetails = CustomUser.objects.get(id=userid)
    if request.method == 'POST':
        formset = PPSRDetailsForm(request.POST)
        if formset.is_valid():
            #print("form valid")
            formset.save()
            return redirect('ppsr_details')  # Replace with your success URL
    else:
        formset = PPSRDetailsForm()
        #print("form not valid")
        userid = request.seesion['userdata']
        #print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
    return render(request, 'Packing/template_name.html', {'formset': formset,'user':userdetails})  # Replace with your template name
@login_required()
def ppsr_details_edit(request):    
    form = PPSRDetailsForm()
    if request.method == 'POST':
        userid = request.session['userdata']
        userdetails = CustomUser.objects.get(id=userid)
        form = PPSRDetailsForm(request.POST)
        if form.is_valid():
            #print("form valid")
            form.save()
            return redirect('ppsrtable')  # Replace with your success URL
    else:
        form = PPSRDetailsForm()
        #print("not post")
        userid = request.session['userdata']
        userdetails = CustomUser.objects.get(id=userid)
    return render(request, 'Packing/ppsrtableentry.html',{'form': form,'user':userdetails})

@method_decorator(login_required,name='dispatch')
class PPSRDetailsListView(ListView):
    model = PPSRDetails
    template_name = 'Packing/ppsrtable.html'
    context_object_name = 'datatable'
    
    breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Production', 'url': '/Packing/productionrolltable'},
        {'label': 'PPSR Details', 'url': None},  # Assuming current page doesn't have a URL
        ] 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = self.breadcrumbs
        
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        context['user'] = userdetails
        
        return context   
class PPSRDetailsCreateView(CreateView):
    model = PPSRDetails
    #form_class = PPSRDetailsFormSet
    template_name = 'Packing/template_name.html'
    success_url = reverse_lazy('ppsrtable')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = PPSRDetailsForm(self.request.POST)
        else:
            context['formset'] = PPSRDetailsForm()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            formset.save()
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))
class PPSRDetailsUpdateView(UpdateView):
    model = PPSRDetails
    form_class = PPSRDetailsForm
    template_name = 'ppsr_details_form.html'
    success_url = reverse_lazy('ppsr_details_list')
class PPSRDetailsDeleteView(DeleteView):
    model = PPSRDetails
    template_name = 'ppsr_details_confirm_delete.html'
    success_url = reverse_lazy('ppsr_details_list')

@method_decorator(login_required,name='dispatch')
class SRDailyStockPETJARUpdateView(View):
    def get(self,request):
        skuname = skunamedetails.objects.filter(godownname=2)
        lastentry = SRDailyStockDetails.objects.last()
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        if lastentry:
            lastupdate = lastentry.date
        else:
            lastupdate = ''      
        form = SRDailyStockDetailsForm(skuname)
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Daily Stock', 'url': '/Packing/stock/dailypetjarlist'},
        {'label': 'Daily Stock Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
        return render(request,'Packing/dailystockpetjartablenewentry.html',{'form':form,'lastentrydate':lastupdate,'user':userdetails,'breadcrumbs':breadcrumbs})
    def post(self,request):
    #     skuname = skunamedetails.objects.filter(
    #     Q(skutype='PET') | Q(skutype='JAR')
    # )
        skuname = skunamedetails.objects.filter(godownname=2)
        form = SRDailyStockDetailsForm(skuname=skuname, data=request.POST)
        updatedby = request.session.get('userdata')
        if updatedby is not None:        
            if form.is_valid():
                stockdata = form.cleaned_data
                stockmode = stockdata.get('stockmode')  # Get the value of stockmode from the cleaned data
                godown = GodownDetails.objects.get(pk=2) # Retrieve the GodownDetails instance with ID 1
                ##print('godown',godown)
                for sku in skuname:
                    # Assuming the correct attribute is 'sku_name'
                    stockbox = stockdata.get(sku.skuname)               
                    
                    if stockbox is not None:
                        SRDailyStockDetails.objects.create(skuname=sku, stockbox=stockbox,stocktype='PET',stockmode=stockmode,updatedby=updatedby,godownname=godown)
                messages.success(request,"PET Stock Details updated Successful")
                return redirect('dailypetjarstocklist')
            else:
                #print('form not valid')
                return render(request, 'Packing/dailystockpetjartablenewentry.html', {'form': form})
        else:
            #print('not updated by')
            messages.error(request,"Unauthenticated Login")
            return render(request, 'Packing/dailystockpetjartablenewentry.html', {'form': form})
@method_decorator(login_required,name='dispatch')
class DailyPETJARstocklist(View):
    def get(self,request): 
        # Step 1: Fetch all relevant records for today's date and specific stock type
        today_date = date.today()
        all_stock_records = SRDailyStockDetails.objects.filter(date__date=today_date, godownname=2)

        # Step 2: Separate the records into morning and evening stock tables
        morningstocktable = [record for record in all_stock_records if record.stockmode == 'Morning']
        eveningstocktable = [record for record in all_stock_records if record.stockmode == 'Evening']
        ##print('morningstocktable',morningstocktable)
        ##print('eveningstocktable',eveningstocktable)
        morningstockerror=0
        eveningstockerror=0
        if not morningstocktable:
            morningstockerror = 1
        sfmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('SF')]
        gnmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GN ')]
        rbmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('RB')]
        cocmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('COC')]
        ginmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GIN')]
        gnrmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GNR')]
        nakrmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('NAK')]
        mustmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('MUS')]
        casmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('CAS')]
        glmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GL')]
        palmmorningstock = [record for record in morningstocktable if 'PALM' in record.skuname.skuname]
        
            
            
        if not eveningstocktable:
            eveningstockerror = 1         
        sfeveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('SF')]
        gneveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GN ')]
        rbeveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('RB')]
        coceveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('COC')]
        gineveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GIN')]
        nakeveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('NAK')]
        musteveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('MUS')]
        caseveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('CAS')]
        gnreveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GNR')]
        gleveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GL')]
        palmeveningstock = [record for record in eveningstocktable if 'PALM' in record.skuname.skuname] 
        
        # #print(eveningstocktable)
        # Step 3: Find the latest date and corresponding updatedby for 'POUCH' stock type
        subquery = SRDailyStockDetails.objects.filter(
            stocktype='PET'
        ).order_by('-date').values('date')[:1]

        latest_record = SRDailyStockDetails.objects.filter(
            stocktype='PET',
            date=Subquery(subquery)
        ).first()

        latestdate = latest_record.date if latest_record else None
        userid = latest_record.updatedby if latest_record else None
        # #print(userid)
        try:
            updatedby = CustomUser.objects.get(id=userid)
        except:
            updatedby = 'unknown'
        
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Daily Stock', 'url': '/Packing/stock/dailypetjarlist'},
        {'label': 'Pet Stock Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
        context = {
            'sfmorningstock':sfmorningstock,
            'gnmorningstock':gnmorningstock,
            'rbmorningstock':rbmorningstock,
            'cocmorningstock':cocmorningstock,
            'ginmorningstock':ginmorningstock,
            'gnrmorningstock':gnrmorningstock,
            'nakmorningstock':nakrmorningstock,
            'mustmorningstock':mustmorningstock,
            'casmorningstock':casmorningstock,
            'glmorningstock':glmorningstock,
            'palmmorningstock':palmmorningstock,
            'eveningstockdata':eveningstocktable,
            'sfeveningstock':sfeveningstock,
            'gneveningstock':gneveningstock,
            'rbeveningstock':rbeveningstock,
            'coceveningstock':coceveningstock,
            'gineveningstock':gineveningstock,
            'gnreveningstock':gnreveningstock,
            'nakeveningstock':nakeveningstock,
            'gleveningstock':gleveningstock,
            'musteveningstock':musteveningstock,
            'caseveningstock':caseveningstock,
            'palmeveningstock':palmeveningstock,
            'lastentrydate':latestdate,
            'updatedby':updatedby,
            'morningstockerror':morningstockerror,
            'eveningstockerror':eveningstockerror,
            'user':userdetails,
            'breadcrumbs':breadcrumbs
        }
        return render(request,'Packing/dailystockpetjartable.html',context)
@method_decorator(login_required,name='dispatch')
class DailystockPouchList(View):
    def get(self,request): 
        # Step 1: Fetch all relevant records for today's date and specific stock type
        today_date = date.today()
        all_stock_records = SRDailyStockDetails.objects.filter(date__date=today_date, godownname=1)

        # Step 2: Separate the records into morning and evening stock tables
        morningstocktable = [record for record in all_stock_records if record.stockmode == 'Morning']
        eveningstocktable = [record for record in all_stock_records if record.stockmode == 'Evening']
        ##print('morningstocktable',morningstocktable)
        ##print('eveningstocktable',eveningstocktable)
        morningstockerror=0
        eveningstockerror=0
        if not morningstocktable:
            morningstockerror = 1
        sfmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('SF')]
        gnmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GN ')]
        rbmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('RB')]
        cocmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('COC')]
        ginmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GIN')]
        gnrmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GNR')]
        glmorningstock = [record for record in morningstocktable if record.skuname.skuname.startswith('GL')]
        palmmorningstock = [record for record in morningstocktable if 'PALM' in record.skuname.skuname]
        
            
            
        if not eveningstocktable:
            eveningstockerror = 1         
        sfeveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('SF')]
        gneveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GN ')]
        rbeveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('RB')]
        coceveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('COC')]
        gineveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GIN')]
        gnreveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GNR')]
        gleveningstock = [record for record in eveningstocktable if record.skuname.skuname.startswith('GL')]
        palmeveningstock = [record for record in eveningstocktable if 'PALM' in record.skuname.skuname] 
        
        # #print(eveningstocktable)
        # Step 3: Find the latest date and corresponding updatedby for 'POUCH' stock type
        subquery = SRDailyStockDetails.objects.filter(
            stocktype='POUCH'
        ).order_by('-date').values('date')[:1]

        latest_record = SRDailyStockDetails.objects.filter(
            stocktype='POUCH',
            date=Subquery(subquery)
        ).first()

        latestdate = latest_record.date if latest_record else None
        userid = latest_record.updatedby if latest_record else None
        # #print(userid)
        try:
            updatedby = CustomUser.objects.get(id=userid)
        except:
            updatedby = 'unknown'
        
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Daily Stock', 'url': '/Packing/stock/dailypouchstocklist'},
        {'label': 'Pouch Stock Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
        context = {
            'sfmorningstock':sfmorningstock,
            'gnmorningstock':gnmorningstock,
            'rbmorningstock':rbmorningstock,
            'cocmorningstock':cocmorningstock,
            'ginmorningstock':ginmorningstock,
            'gnrmorningstock':gnrmorningstock,
            'glmorningstock':glmorningstock,
            'palmmorningstock':palmmorningstock,
            'eveningstockdata':eveningstocktable,
            'sfeveningstock':sfeveningstock,
            'gneveningstock':gneveningstock,
            'rbeveningstock':rbeveningstock,
            'coceveningstock':coceveningstock,
            'gineveningstock':gineveningstock,
            'gnreveningstock':gnreveningstock,
            'gleveningstock':gleveningstock,
            'palmeveningstock':palmeveningstock,
            'lastentrydate':latestdate,
            'updatedby':updatedby,
            'morningstockerror':morningstockerror,
            'eveningstockerror':eveningstockerror,
            'user':userdetails,
            'breadcrumbs':breadcrumbs
        }
        return render(request,'Packing/dailystockpouchtable.html',context)
@method_decorator(login_required,name='dispatch')   
class DailystockPouchUpdate(View):
    def get(self,request):
        skuname = skunamedetails.objects.filter(godownname=1)
        lastentry = SRDailyStockDetails.objects.last()
        if lastentry:
            lastupdate = lastentry.date
        else:
            lastupdate = ''      
        form = SRDailyStockDetailsForm(skuname)
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Daily Stock', 'url': '/Packing/stock/dailypouchstocklist'},
        {'label': 'Pouch Stock Update', 'url': None},  # Assuming current page doesn't have a URL
        ]
        return render(request,'Packing/dailystockpouchtablenewentry.html',{'form':form,'lastentrydate':lastupdate,'user':userdetails,'breadcrumbs':breadcrumbs})
    
    def post(self,request):
        skuname = skunamedetails.objects.filter(skutype='POUCH')
        form = SRDailyStockDetailsForm(skuname=skuname, data=request.POST)
        updatedby = request.session.get('userdata')
        if updatedby is not None:        
            if form.is_valid():
                stockdata = form.cleaned_data
                stockmode = stockdata.get('stockmode')  # Get the value of stockmode from the cleaned data
                godown = GodownDetails.objects.get(pk=1) # Retrieve the GodownDetails instance with ID 1
                for sku in skuname:
                    # Assuming the correct attribute is 'sku_name'
                    stockbox = stockdata.get(sku.skuname)               
                    
                    if stockbox is not None:
                        SRDailyStockDetails.objects.create(skuname=sku, stockbox=stockbox,stocktype='POUCH',stockmode=stockmode,updatedby=updatedby,godownname=godown)
                messages.success(request,"Pouch Stock Details updated Successful")
                return redirect('dailypouchstocklist')
            else:
                #print('form not valid')
                return render(request, 'Packing/dailystockpouchtablenewentry.html', {'form': form})
        else:
            #print('not updated by')
            messages.error(request,"Unauthenticated Login")
            return render(request, 'Packing/dailystockpouchtablenewentry.html', {'form': form})
@method_decorator(login_required,name='dispatch')
class DailyStockSRFull(View):
    def get(self,request):
        today_date = date.today()
        dayfullstock = SRDailyStockDetails.objects.filter(date__date=today_date)
        morningstocktable_pouchgodown = [record for record in dayfullstock if record.stockmode == 'Morning' and record.godownname.godownname == "SR-Packing"]
        morningstocktable_jargodown = [record for record in dayfullstock if record.stockmode == 'Morning' and record.godownname.godownname == "SR-Jar"]
        eveningstocktable_pouchgodown = [record for record in dayfullstock if record.stockmode == 'Evening' and record.godownname.godownname == "SR-Packing"]
        eveningstocktable_jargodown = [record for record in dayfullstock if record.stockmode == 'Evening' and record.godownname.godownname == "SR-Jar"]
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Exp Vs Act', 'url': '/Packing/stock/dailystockSRFull'},
        {'label': 'Exp Vs Act Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
        context = {
            'morningpochfullstock' : morningstocktable_pouchgodown,
            'morningpetjarfullstock' : morningstocktable_jargodown,
            'eveningpochfullstock' : eveningstocktable_pouchgodown,
            'eveningpetjarfullstock' : eveningstocktable_jargodown,
            'user':userdetails,
            'bradcrumbs':breadcrumbs
        }
        
        return render(request, 'Packing/dailystock_sr_stock_full.html', context)    

@method_decorator(login_required,name='dispatch')
class DispatchReqVsStockView(View):
    def get(self,request):
        if 'download' in request.GET:
            return self.download_excel(request)
        else:
            return self.showdata(request)
    def showdata(self,request):
        today=date.today()
        context = {
            'error': None
        }
        eveningfullstock = SRDailyStockDetails.objects.filter(date__date=today,stockmode='Evening')
        # #print(eveningfullstock)
        dispatcheveingreq = DispatchReq.objects.filter(date=today)
        # #print(dispatcheveingreq)
        if eveningfullstock:
            if dispatcheveingreq:
                stockdf = pd.DataFrame(list(eveningfullstock.values('skuname','stockbox')))
                # #print(stockdf)
                reqdf = pd.DataFrame(list(dispatcheveingreq.values('skuname','reqbox')))
                # #print(reqdf)
                
                # mergedf = pd.merge(stockdf,reqdf,how='outer')
                # #print(mergedf)
                # #print('InnerJoint')
                # mergedf = pd.merge(stockdf,reqdf,how='inner')
                mergedf = pd.merge(stockdf,reqdf,how='right')
                # #print(mergedf)
                
                mergedf.fillna(0,inplace=True)
                # #print('fil0 : ',mergedf)
                
                mergedf['stockavail']=mergedf['stockbox']-mergedf['reqbox']
                # #print('calc',mergedf)
                
                plandata = mergedf.to_dict('records')
                # #print(plandata)
                context['plandata'] = plandata
                
                # Replace skuname IDs with actual skuname names
                skunames = {skuname.id: skuname.skuname for skuname in skunamedetails.objects.all()}
                for item in plandata:
                    item['skuname'] = skunames.get(item['skuname'], 'Unknown')
                    # #print(item['skuname'])
            else:
                context['error'] = "Dispatch requirement not updated"
        else:
            context['error'] = 'Evening Stock not updated' 
        breadcrumbs = [
        {'label': 'Home', 'url': '/Packing'},
        {'label': 'Dispatch Req Vs Stock', 'url': '/Packing/MIS/dispatchreqvsstockview'},
        {'label': 'Req Vs Stock Table', 'url': None},  # Assuming current page doesn't have a URL
        ]
        context['breadcrumbs'] = breadcrumbs   
        return render(request,'Packing/dispatchreqvsstockview.html',context)
    def download_excel(self,request):
        today=date.today()
        context = {
            'error': None
        }
        eveningfullstock = SRDailyStockDetails.objects.filter(date__date=today,stockmode='Evening')
        # #print(eveningfullstock)
        dispatcheveingreq = DispatchReq.objects.filter(date=today)
        # #print(dispatcheveingreq)
        if eveningfullstock:
            if dispatcheveingreq:
                stockdf = pd.DataFrame(list(eveningfullstock.values('skuname','stockbox')))
                # #print(stockdf)
                reqdf = pd.DataFrame(list(dispatcheveingreq.values('skuname','reqbox')))
                # #print(reqdf)
                
                # mergedf = pd.merge(stockdf,reqdf,how='outer')
                # #print(mergedf)
                # #print('InnerJoint')
                # mergedf = pd.merge(stockdf,reqdf,how='inner')
                mergedf = pd.merge(stockdf,reqdf,how='right')
                # #print(mergedf)
                
                mergedf.fillna(0,inplace=True)
                # #print('fil0 : ',mergedf)
                
                mergedf['stockavail']=mergedf['stockbox']-mergedf['reqbox']
                # #print('calc',mergedf)
                
                # Replace skuname IDs with actual skuname names
                skunames = {skuname.id: skuname.skuname for skuname in skunamedetails.objects.all()}
                mergedf['skuname'] = mergedf['skuname'].map(skunames)
            else:
                context['error'] = "Dispatch requirement not updated"
                mergedf = pd.DataFrame(columns=['skuname', 'stockbox', 'reqbox', 'stockavail'])
        else:
            context['error'] = 'Evening Stock not updated'
            mergedf = pd.DataFrame(columns=['skuname', 'stockbox', 'reqbox', 'stockavail']) 
        # Generate Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dispatch_ReqvsStock_{today}.xlsx"'
        
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            mergedf.to_excel(writer, index=False, sheet_name='Dispatch_ReqVsStock')

        return response             
@method_decorator(login_required,name='dispatch')
class DispatchReqVsStockUpdate(View):
    def get(self,request):
        form = DispatchReqForm()
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        
        return render(request, 'Packing/dispatchreqform.html', {'form': form,'user':userdetails})
        
    def post(self,request):
        forms = []
        userid = self.request.session.get('userdata')
        ##print('userid',userid)
        userdetails = CustomUser.objects.get(id=userid)
        for i in range(len(request.POST.getlist('skuname'))):
            form = DispatchReqForm({
                'skuname': request.POST.getlist('skuname')[i],
                'reqbox': request.POST.getlist('reqbox')[i]
            })
            if form.is_valid():
                DispatchReq.objects.create(
                    skuname=form.cleaned_data['skuname'],
                    reqbox=form.cleaned_data['reqbox']
                )
        messages.success(request,"Dispatch Req saved Successful.")
        form = DispatchReqForm()
        return render(request, 'Packing/dispatchreqform.html', {'form': form,'user':userdetails})  # Redirect to a success page or another view

class CalenderView(View):
    def get(self,request):
        return render(request,'Packing/calendar.html')
# class GenericListView(ListView):
#     template_name = 'Packing/generictemplate.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['view'] = {
#             'action': 'list',
#             'create_url': reverse_lazy(f'{self.model._meta.model_name}-create'),
#             'model_name': self.model._meta.verbose_name.capitalize(),
#         }
#         return context

# class GenericCreateView(CreateView):
#     template_name = 'Packing/generic_template.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['view'] = {
#             'action': 'form',
#             'submit_label': 'Create',
#         }
#         return context

# class GenericUpdateView(UpdateView):
#     template_name = 'Packing/generic_template.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['view'] = {
#             'action': 'form',
#             'submit_label': 'Update',
#         }
#         return context

# class GenericDeleteView(DeleteView):
#     template_name = 'Packing/generic_template.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['view'] = {
#             'action': 'delete',
#             'cancel_url': reverse_lazy(f'{self.model._meta.model_name}-list'),
#         }
#         return context

#     def get_success_url(self):
#         return reverse_lazy(f'{self.model._meta.model_name}-list')

# class BrandListView(GenericListView):
#     model = branddetails

# class BrandCreateView(GenericCreateView):
#     model = branddetails
#     fields = ['brandname']

# class BrandUpdateView(GenericUpdateView):
#     model = branddetails
#     fields = ['brandname']

# class BrandDeleteView(GenericDeleteView):
#     model = branddetails

# class OilCategoryListView(GenericListView):
#     model = oilcategorydetails

# class OilCategoryCreateView(GenericCreateView):
#     model = oilcategorydetails
#     fields = ['brandname', 'oilcategoryname']

# class OilCategoryUpdateView(GenericUpdateView):
#     model = oilcategorydetails
#     fields = ['brandname', 'oilcategoryname']

# class OilCategoryDeleteView(GenericDeleteView):
#     model = oilcategorydetails

# class SkuNameListView(GenericListView):
#     model = skunamedetails

# class SkuNameCreateView(GenericCreateView):
#     model = skunamedetails
#     fields = ['oilcategoryname', 'skuname', 'skucode']

# class SkuNameUpdateView(GenericUpdateView):
#     model = skunamedetails
#     fields = ['oilcategoryname', 'skuname', 'skucode']

# class SkuNameDeleteView(GenericDeleteView):
#     model = skunamedetails

# Create your views here.