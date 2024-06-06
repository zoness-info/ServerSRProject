# myapp/management/commands/import_excel_data.py
# Run below comment to write db from excel
#python manage.py import_excel_data /path/to/your/excel/file.xlsx

#python manage.py importexceldata /path/to/your/excel/file.xlsx SheetName
#python manage.py import_excel_data E:\datafiles\masterdata.xlsx


#file_path = "E:\datafiles\masterdata.xlsx"

import os
import openpyxl
from django.core.management.base import BaseCommand
from Packing.models import (
                    branddetails, oilcategorydetails, skunamedetails, 
                             DayNightshift, 
                            FilmRollType, OperatorNameDetails, PackingMachineDetails, 
                            PackingSection,
                            MainTankDetails,SubTankDetails,VitaminDetails,TBHQDetails,TMPSDetails,QCNameDetails,
                            PouchLeakMistakesName,GodownDetails
)

class Command(BaseCommand):
    help = 'Import data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File '{file_path}' does not exist"))
            return

        try:
            wb = openpyxl.load_workbook(file_path)

            # Process each sheet for the corresponding model
            self.import_brand_details(wb['brandname'])
            self.import_oil_category_details(wb['oilcategory'])
            self.import_godown_details(wb['godown'])
            self.import_sku_name_details(wb['sku'])
            
            
            self.import_day_night_shift(wb['shift'])
            self.import_packing_section(wb['packingsection'])
            self.import_operator_name_details(wb['operatorname'])            
            
            self.import_packing_machine_details(wb['packmachine'])
            self.import_maintank(wb['maintank'])
            self.import_subtank(wb['subtank'])
            self.import_vitamin(wb['vitamin'])
            self.import_tmps(wb['tmps'])
            self.import_tbhq(wb['tbhq'])
            self.import_qcname(wb['qcname'])  
            self.import_pouchmistakename(wb['mistakename'])
            

            self.stdout.write(self.style.SUCCESS('All Data imported successfully'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"An error occurred: {str(e)}"))

    def import_brand_details(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            branddetails.objects.create(brandname=row[0])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Brand Name imported successfully'))
        
    def import_godown_details(self, sheet):
        self.stdout.write(self.style.SUCCESS('Godown Details Start'))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            GodownDetails.objects.create(godownname=row[0])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Godown Details imported successfully'))

    def import_day_night_shift(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            DayNightshift.objects.create(shifttype=row[0])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Shift Details imported successfully'))
        
    def import_packing_section(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            PackingSection.objects.create(sectionname=row[0])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Packing Section imported successfully'))    
        
        
    def import_operator_name_details(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sectionname_str = row[1]
            sectionname, createdat = PackingSection.objects.get_or_create(sectionname=sectionname_str)
            OperatorNameDetails.objects.create(empid=row[0],sectionname=sectionname,operatorname=row[2],mobileno=row[3])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Operator Details imported successfully'))
    
    def import_oil_category_details(self, sheet):
        self.stdout.write(self.style.SUCCESS('Oil Category Details Start'))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            brandname_str = row[0]
            brandname, createdat = branddetails.objects.get_or_create(brandname=brandname_str)
            oilcategorydetails.objects.create(brandname=brandname,oilcategoryname=row[1])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Oil Category Details imported successfully'))
        
    def import_sku_name_details(self, sheet):
        self.stdout.write(self.style.SUCCESS('SKU Details Start'))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            #print('rowno:', row)
            category_id_str = row[0]
            godownname_str = row[5] 
            # print(category_id_str)

            category, created = oilcategorydetails.objects.get_or_create(oilcategoryname=category_id_str)
            godownname, created = GodownDetails.objects.get_or_create(godownname = godownname_str)
            #print(category.id,category.brandname,category.oilcategoryname)                
            skunamedetails.objects.create(category_name=category,skuname=row[1],skutype=row[2],skucode_m=row[3],skucode_c=row[4],godownname=godownname)  # Adjust fields accordingly        
        self.stdout.write(self.style.SUCCESS('SKU Details imported successfully'))
        
    def import_packing_machine_details(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            PackingMachineDetails.objects.create(machineid=row[0],machinename=row[1])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Packing Machine Details imported successfully'))   
        
    def import_maintank(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            MainTankDetails.objects.create(maintankname=row[0],oilname=row[1],desc=row[2],capacity=row[3])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Main Tank Details imported successfully'))
        
    def import_subtank(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            SubTankDetails.objects.create(subtankname=row[0],oilname=row[1],desc=row[2],capacity=row[3])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('SubTank Details imported successfully'))
        
    def import_vitamin(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            VitaminDetails.objects.create(vitaminname=row[0],units=row[1],desc=row[2])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Vitamin Details imported successfully'))
    
    def import_tmps(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            TMPSDetails.objects.create(units=row[0],desc=row[1])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('TMPS Details imported successfully'))
    
    def import_tbhq(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            TBHQDetails.objects.create(units=row[0],desc=row[1])  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('TBHQ Details imported successfully'))
    
    def import_qcname(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            QCNameDetails.objects.create(qcname=row[0],)  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('QC Name Details imported successfully')) 
    
    def import_pouchmistakename(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            PouchLeakMistakesName.objects.create(mistakename=row[0],)  # Adjust fields accordingly
        self.stdout.write(self.style.SUCCESS('Pouch Mistake Details imported successfully'))